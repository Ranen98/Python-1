from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    # 打开文件
    file = load_workbook(path)
    # 所有表格的名称
    sheets = file.get_sheet_names()

    # 拿出一个表格
    sheet = file.get_sheet_by_name(sheets[1])
    for lineNum in range(1, sheet.max_row + 1):
        # print(lineNum)
        lineList = []
        # print(sheet.max_row, sheet.max_column)
        for columnNum in range(1, sheet.max_column + 1):
            # 拿数据
            value = sheet.cell(row=lineNum, column=columnNum).value
            if value != None:
                lineList.append(value)
            else:
                lineList.append("无")

        print(lineList)

# 不能读取xls类型的文件
path = r"C:\学习\Python语言基础\15、2、自动化办公与鼠标键盘模拟\5、excel自动化办公\sunck.xlsx"
readXlsxFile(path)